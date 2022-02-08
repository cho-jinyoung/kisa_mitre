from selenium import webdriver
from bs4 import BeautifulSoup
import chromedriver_autoinstaller
import time
import openpyxl
from openpyxl import load_workbook
    
wb=openpyxl.Workbook()  # 작업할 workbook생성
sheet=wb.active         # 작업할 workbook내 sheet활성화

read = load_workbook("MITRE technique-discription.xlsx", data_only=True)    # 기존 파일 읽기
read_sheet=read['MITRE ATT&CK v10.0']                                       # 파일 내 시트명
read_cell=read_sheet['A2:A568']

# 웹사이트 열기
chrome_ver = chromedriver_autoinstaller.get_chrome_version().split('.')[0]

try:
    driver=webdriver.Chrome(f"./{chrome_ver}./chromedriver.exe")
except:
    chromedriver_autoinstaller.install(True)
    driver = webdriver.Chrome(f'./{chrome_ver}/chromedriver.exe')
    
for row in read_cell:   # 행
    for cell in row:    # 열
        sheet.append([cell.value])
        
        ipt=cell.value
        ipt=ipt.replace(".", "/")
        url="https://attack.mitre.org/techniques/"+ipt+"/"
        driver.get(url) 

        time.sleep(3)           # selenium이 웹페이지를 읽어올 시간을 줌
        html=driver.page_source
        
        for page in range(1, 3):
            try:
                id_miti=driver.find_element_by_xpath('//*[@id="v-attckmatrix"]/div[2]/div/div/div/table[%s]/thead/tr/th[2]' %page)
                if id_miti.text == 'Mitigation':
                    thread=driver.find_element_by_xpath('//*[@id="v-attckmatrix"]/div[2]/div/div/div/table[%s]/tbody' %page)
                    tr=thread.find_elements_by_tag_name('tr')
                    for i in tr:
                        sheet.append([cell.value, i.text[:5], i.text[5:]])
            except:
                pass    

wb.save("tec_miti.xlsx")
